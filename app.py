import os
from io import BytesIO
import pandas as pd
import streamlit as st
import base64

st.set_page_config(
    page_title="Frequency Estimator",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="collapsed",
)


def check_password():
    def password_entered():
        if st.session_state["password"] == st.secrets["APP_PASSWORD"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("Enter password", type="password", key="password", on_change=password_entered)
        st.stop()

    elif not st.session_state["password_correct"]:
        st.text_input("Enter password", type="password", key="password", on_change=password_entered)
        st.error("Incorrect password")
        st.stop()

check_password()


BRAND_COLOR = "#F48020"
BRAND_DARK = "#d96d15"
BG_LIGHT = "#fff7f1"
TEXT_DARK = "#1f2937"
TEXT_MUTED = "#6b7280"
BORDER = "#f3d2ba"

FACTORS = [
    {"section": "Marketing", "factor": "Brand Proposition", "low": "Established", "high": "New",
     "weight": 4.545454545454546, "sample": 6},
    {"section": "Marketing", "factor": "Market Share", "low": "High", "high": "Low", "weight": 4.545454545454546,
     "sample": 4},
    {"section": "Marketing", "factor": "Brand Loyalty", "low": "High", "high": "Low", "weight": 4.545454545454546,
     "sample": 4},
    {"section": "Marketing", "factor": "Purchase Cycle", "low": "Long", "high": "Short", "weight": 4.545454545454546,
     "sample": 4},
    {"section": "Marketing", "factor": "Frequency of Usage", "low": "Less", "high": "More", "weight": 4.545454545454546,
     "sample": 6},
    {"section": "Marketing", "factor": "Age of Audience", "low": "Young", "high": "Old", "weight": 4.545454545454546,
     "sample": 4},
    {"section": "Marketing", "factor": "Competitive Activity (SOV)", "low": "Low", "high": "High",
     "weight": 4.545454545454546, "sample": 6},
    {"section": "Marketing", "factor": "Habits / Attitude", "low": "Reinforce", "high": "Change",
     "weight": 4.545454545454546, "sample": 6},
    {"section": "Marketing", "factor": "Competitive Threat", "low": "Low", "high": "High", "weight": 4.545454545454546,
     "sample": 6},

    {"section": "Creative", "factor": "Message Complexity", "low": "Simple", "high": "Complex",
     "weight": 4.545454545454546, "sample": 1},
    {"section": "Creative", "factor": "Proposition (Message Uniqueness)", "low": "Established", "high": "New",
     "weight": 4.545454545454546, "sample": 6},
    {"section": "Creative", "factor": "Commercial (Continuing Campaign?)", "low": "Established", "high": "New",
     "weight": 4.545454545454546, "sample": 7},
    {"section": "Creative", "factor": "Message Skew", "low": "Product", "high": "Image", "weight": 4.545454545454546,
     "sample": 7},
    {"section": "Creative", "factor": "Message Variety (No. of Creatives)", "low": "Low", "high": "High",
     "weight": 4.545454545454546, "sample": 1},
    {"section": "Creative", "factor": "Wearout (Freshness Retention)", "low": "High", "high": "Low",
     "weight": 4.545454545454546, "sample": 3},
    {"section": "Creative", "factor": "Size of Ad Unit", "low": "Large", "high": "Small", "weight": 4.545454545454546,
     "sample": 5},

    {"section": "Media", "factor": "Ad Clutter", "low": "Low", "high": "High", "weight": 4.545454545454546,
     "sample": 6},
    {"section": "Media", "factor": "Ad Environment", "low": "Compatible", "high": "Incompatible",
     "weight": 4.545454545454546, "sample": 6},
    {"section": "Media", "factor": "Audience Attentiveness", "low": "High", "high": "Low", "weight": 4.545454545454546,
     "sample": 4},
    {"section": "Media", "factor": "Scheduling", "low": "Continuous", "high": "Flighting", "weight": 4.545454545454546,
     "sample": 4},
    {"section": "Media", "factor": "Repeat Exposure Media", "low": "High", "high": "Low", "weight": 4.545454545454546,
     "sample": 2},
    {"section": "Media", "factor": "Media Mix", "low": "Mix", "high": "Single", "weight": 4.545454545454546,
     "sample": 2},
]

SECTIONS = ["Marketing", "Creative", "Media"]
SCALE = list(range(1, 8))
LOGO_PATH = os.path.join("assets", "logo_freq.png")


def ensure_state(mode: str = "neutral"):
    for i, item in enumerate(FACTORS):
        key = f"score_{i}"
        if key not in st.session_state:
            st.session_state[key] = 4 if mode == "neutral" else item["sample"]

    if "generated" not in st.session_state:
        st.session_state.generated = False

    if "result_payload" not in st.session_state:
        st.session_state.result_payload = None


def reset_scores(mode: str = "neutral"):
    for i, item in enumerate(FACTORS):
        st.session_state[f"score_{i}"] = 4 if mode == "neutral" else item["sample"]

    st.session_state.generated = False
    st.session_state.result_payload = None


def calculate_results() -> tuple[float, pd.DataFrame, pd.DataFrame]:
    records = []
    for i, item in enumerate(FACTORS):
        score = st.session_state[f"score_{i}"]
        records.append(
            {
                **item,
                "score": score,
                "weighted_score": score * item["weight"],
            }
        )

    df = pd.DataFrame(records)

    efficient_frequency = round(df["weighted_score"].sum() / df["weight"].sum(), 2)

    section_summary = (
        df.groupby("section", as_index=False)
        .agg(avg_score=("score", "mean"), factors=("factor", "count"))
    )
    section_summary["avg_score"] = section_summary["avg_score"].round(2)

    return efficient_frequency, df, section_summary


def interpret_score(score: float) -> tuple[str, str]:
    if score < 3:
        return "Low", "The brand may require relatively lower effective frequency pressure."
    if score < 5:
        return "Moderate", "A balanced effective frequency level is likely appropriate."
    return "High", "The brand may benefit from higher effective frequency pressure."


def generate_result():
    score, detail_df, section_df = calculate_results()
    label, note = interpret_score(score)

    st.session_state.result_payload = {
        "score": score,
        "detail_df": detail_df,
        "section_df": section_df,
        "label": label,
        "note": note,
    }
    st.session_state.generated = True


def build_export_file(
        detail_df: pd.DataFrame,
        section_df: pd.DataFrame,
        score: float,
        label: str,
        note: str,
) -> BytesIO:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    export_detail = detail_df[["section", "factor", "low", "high", "score"]].copy()
    export_detail.columns = [
        "Section",
        "Factor",
        "Score 1 Meaning",
        "Score 7 Meaning",
        "Selected Score",
    ]

    summary_df = pd.DataFrame(
        [
            {"Metric": "Desired Efficient Frequency", "Value": score},
            {"Metric": "Assessment", "Value": label},
            {"Metric": "Comment", "Value": note},
        ]
    )

    export_sections = section_df.copy()
    export_sections.columns = ["Section", "Average Score", "No. of Factors"]

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Result Summary", index=False)
        export_sections.to_excel(writer, sheet_name="Section Summary", index=False)
        export_detail.to_excel(writer, sheet_name="Scoring Detail", index=False)

        header_fill = PatternFill(fill_type="solid", fgColor=BRAND_COLOR.replace("#", ""))
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )

        for sheet_name, df in {
            "Result Summary": summary_df,
            "Section Summary": export_sections,
            "Scoring Detail": export_detail,
        }.items():
            ws = writer.sheets[sheet_name]

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            for idx, col in enumerate(df.columns, start=1):
                max_len = max(len(str(col)), *(len(str(v)) for v in df[col].fillna("")))
                adjusted_width = min(max(max_len + 3, 16), 45)
                ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    output.seek(0)
    return output


ensure_state()

st.markdown(
    f"""
    <style>
    .stApp {{
        background: linear-gradient(180deg, #fffaf6 0%, #fff3eb 100%);
    }}

    .block-container {{
        max-width: 1320px;
        padding-top: 0.2rem;
        padding-bottom: 0.4rem;
    }}

    .header-wrap {{
        background: white;
        border: 1px solid {BORDER};
        border-radius: 6px;
        padding: 0.35rem 0.5rem;
        box-shadow: none;
        margin-bottom: 0.45rem;
    }}

    .header-row {{
        display: flex;
        align-items: stretch;
        gap: 0.75rem;
    }}

    .logo-box {{
        width: 120px;
        height: 78px;
        display: flex;
        align-items: center;
        justify-content: center;
        background: white;
        border-radius: 4px;
        overflow: hidden;
    }}

    .logo-box img {{
        max-height: 90px;
        max-width: 100%;
        object-fit: contain;
        display: block;
    }}

    .hero-box {{
        flex: 1;
        height: 78px;
        background: linear-gradient(135deg, {BRAND_COLOR} 0%, {BRAND_DARK} 100%);
        color: white;
        border-radius: 4px;
        padding: 0.7rem 1rem;
        display: flex;
        flex-direction: column;
        justify-content: center;
        box-sizing: border-box;
    }}

    .hero-title {{
        font-size: 1.85rem;
        font-weight: 700;
        margin: 0;
        line-height: 1.05;
    }}

    .hero-subtitle {{
        font-size: 0.92rem;
        margin: 0.18rem 0 0 0;
        opacity: 0.95;
        line-height: 1.2;
    }}

    .toolbar-note {{
        color: {TEXT_MUTED};
        font-size: 0.88rem;
        padding-top: 0.15rem;
        margin-top: 0;
    }}

    .section-card {{
        background: white;
        border: 1px solid {BORDER};
        border-radius: 4px;
        padding: 0.9rem 1rem 0.55rem 1rem;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03);
        margin-bottom: 0.8rem;
    }}

    .summary-card {{
        background: white;
        border: 1px solid {BORDER};
        border-radius: 10px;
        padding: 1rem;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03);
        position: sticky;
        top: 1rem;
    }}

    .factor-label {{
        font-weight: 600;
        color: {TEXT_DARK};
        font-size: 0.96rem;
        line-height: 1.35;
    }}

    .range-pill {{
        display: inline-block;
        font-size: 0.78rem;
        color: {TEXT_DARK};
        background: {BG_LIGHT};
        border: 1px solid {BORDER};
        border-radius: 999px;
        padding: 0.32rem 0.7rem;
        white-space: nowrap;
    }}

    .section-title {{
        color: {BRAND_DARK};
        font-weight: 700;
        margin-bottom: 0.7rem;
    }}

    div[data-testid="stMetric"] {{
        background: #fffaf6;
        border: 1px solid {BORDER};
        border-radius: 4px;
        padding: 0.7rem 0.85rem;
    }}

    div[data-testid="stMetricLabel"] {{
        color: {TEXT_MUTED};
        font-weight: 600;
    }}

    div[data-testid="stMetricValue"] {{
        color: {TEXT_DARK};
    }}

    .stButton > button {{
        border-radius: 10px;
        font-weight: 600;
        padding-top: 0.5rem;
        padding-bottom: 0.5rem;
    }}

    .stDownloadButton > button {{
        border-radius: 10px;
        font-weight: 600;
        padding-top: 0.5rem;
        padding-bottom: 0.5rem;
        width: 100%;
    }}

    .stAlert {{
        border-radius: 10px;
    }}

    header[data-testid="stHeader"] {{
        display: none;
    }}

    div[data-testid="stToolbar"] {{
        display: none;
    }}

    .block-container {{
        padding-top: 0rem !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)




def get_base64_image(image_path: str) -> str:
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode()


if os.path.exists(LOGO_PATH):
    logo_b64 = get_base64_image(LOGO_PATH)
    logo_html = f'<img src="data:image/png;base64,{logo_b64}" alt="Logo">'
else:
    logo_html = f"""
    <div style="
        width:100%;
        height:100%;
        display:flex;
        align-items:center;
        justify-content:center;
        color:{TEXT_MUTED};
        font-size:0.9rem;
        border:1px dashed {BORDER};
        border-radius:4px;">
        Logo not found
    </div>
    """

st.markdown(
    f"""
    <div class="header-wrap">
        <div class="header-row">
            <div class="logo-box">
                {logo_html}
            </div>
            <div class="hero-box">
                <div class="hero-title">Frequency Estimator</div>
                <div class="hero-subtitle">
                    Adapted from the Ostrow Model, with customizable factors based on campaign needs
                </div>
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

ctrl1, ctrl2, ctrl3 = st.columns([1.2, 1.35, 4.2])

with ctrl1:
    if st.button("Reset to neutral", use_container_width=True):
        reset_scores("neutral")

with ctrl2:
    if st.button("Generate result", type="primary", use_container_width=True):
        generate_result()

with ctrl3:
    st.markdown(
        '<div class="</div>',
        unsafe_allow_html=True,
    )

left, right = st.columns([1.9, 1], gap="large")

with left:
    for section in SECTIONS:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown(f'<div class="section-title">{section}</div>', unsafe_allow_html=True)

        section_items = [item for item in FACTORS if item["section"] == section]

        for item in section_items:
            idx = FACTORS.index(item)

            c1, c2, c3 = st.columns([1.55, 2.6, 1.5], vertical_alignment="center")

            with c1:
                st.markdown(
                    f'<div class="factor-label">{item["factor"]}</div>',
                    unsafe_allow_html=True,
                )

            with c2:
                st.select_slider(
                    label=item["factor"],
                    options=SCALE,
                    key=f"score_{idx}",
                    label_visibility="collapsed",
                )

            with c3:
                st.markdown(
                    f'<div class="range-pill">1 = {item["low"]} &nbsp;&nbsp;|&nbsp;&nbsp; 7 = {item["high"]}</div>',
                    unsafe_allow_html=True,
                )

        st.markdown('</div>', unsafe_allow_html=True)

with right:
    st.markdown('<div class="summary-card">', unsafe_allow_html=True)
    st.subheader("Result")

    if st.session_state.generated and st.session_state.result_payload:
        payload = st.session_state.result_payload
        score = payload["score"]
        section_df = payload["section_df"]
        detail_df = payload["detail_df"]
        label = payload["label"]
        note = payload["note"]

        st.metric("Desired Efficient Frequency", f"{score:.2f}")
        st.metric("Assessment", label)
        st.caption(note)

        progress_value = min(max((score - 1) / 6, 0), 1)
        st.progress(progress_value, text=f"Score position on 1–7 scale: {score:.2f}")

        # st.markdown("#### Section averages")
        # chart_df = section_df.set_index("section")[["avg_score"]]
        # st.bar_chart(chart_df)

        export_file = build_export_file(detail_df, section_df, score, label, note)

        st.download_button(
            "Download result in Excel",
            data=export_file,
            file_name="frequency_estimator_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with st.expander("Scoring summary", expanded=False):
            view_df = detail_df[["section", "factor", "score"]].copy()
            view_df["score"] = view_df["score"].astype(int)
            st.dataframe(view_df, use_container_width=True, hide_index=True)
    else:
        st.info("Result will appear here after you click Generate result.")

    st.markdown('</div>', unsafe_allow_html=True)

st.caption("Joseph W. Ostrow - Frequency Estimator Model 1982")